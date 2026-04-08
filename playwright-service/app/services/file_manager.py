"""
File Manager - Handles file operations for downloaded reports.

Provides utilities for:
- File naming with timestamps
- File validation
- Cleanup of old files
- Excel file consolidation (supports multi-sheet source files with custom sheet name mapping)
"""

import os
import shutil
import urllib.parse
from datetime import datetime
from typing import Optional, List
import structlog
from openpyxl import load_workbook, Workbook

from app.config import get_settings

logger = structlog.get_logger()
settings = get_settings()


class FileManager:
    """Manages file operations for downloaded reports."""

    def __init__(self):
        self.download_dir = settings.download_dir
        self.screenshot_dir = settings.screenshot_dir
        self._ensure_directories()

    def _ensure_directories(self) -> None:
        """Ensure required directories exist."""
        os.makedirs(self.download_dir, exist_ok=True)
        os.makedirs(self.screenshot_dir, exist_ok=True)

    def generate_filename(
        self,
        report_type: str,
        tenant_name: str,
        period: Optional[str] = None,
        extension: str = "xlsx"
    ) -> str:
        """Generate a standardized filename for a report."""
        safe_tenant = self._sanitize_filename(tenant_name)
        report_name = report_type.replace("_", " ").title().replace(" ", "_")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if period:
            filename = f"{report_name}_{safe_tenant}_{period}_{timestamp}.{extension}"
        else:
            filename = f"{report_name}_{safe_tenant}_{timestamp}.{extension}"

        return filename

    def _sanitize_filename(self, name: str) -> str:
        """Sanitize a string for use in filenames."""
        invalid_chars = '<>:"/\\|?*'
        result = name
        for char in invalid_chars:
            result = result.replace(char, '')
        result = result.replace(' ', '_')
        while '__' in result:
            result = result.replace('__', '_')
        return result[:100]

    def rename_download(self, original_path: str, new_filename: str) -> str:
        """Rename a downloaded file."""
        if not os.path.exists(original_path):
            raise FileNotFoundError(f"File not found: {original_path}")

        new_path = os.path.join(self.download_dir, new_filename)
        if os.path.exists(new_path):
            os.remove(new_path)

        shutil.move(original_path, new_path)
        return new_path

    def get_file_info(self, filepath: str) -> dict:
        """Get information about a file."""
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"File not found: {filepath}")

        stat = os.stat(filepath)
        return {
            "path": filepath,
            "filename": os.path.basename(filepath),
            "size": stat.st_size,
            "created_at": datetime.fromtimestamp(stat.st_ctime).isoformat(),
            "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        }

    def list_downloads(self) -> list[dict]:
        """List all files in the download directory."""
        files = []
        for filename in os.listdir(self.download_dir):
            filepath = os.path.join(self.download_dir, filename)
            if os.path.isfile(filepath):
                try:
                    files.append(self.get_file_info(filepath))
                except Exception:
                    pass
        return sorted(files, key=lambda x: x["modified_at"], reverse=True)

    def cleanup_old_files(self, max_age_days: int = 30) -> int:
        """Remove files older than specified age."""
        from datetime import timedelta

        cutoff = datetime.now() - timedelta(days=max_age_days)
        deleted = 0

        for filename in os.listdir(self.download_dir):
            filepath = os.path.join(self.download_dir, filename)
            if os.path.isfile(filepath):
                mtime = datetime.fromtimestamp(os.stat(filepath).st_mtime)
                if mtime < cutoff:
                    os.remove(filepath)
                    deleted += 1

        return deleted

    def copy_to_onedrive(
        self,
        source_path: str,
        onedrive_origin: str,
        client_onedrive_folder: str,
    ) -> str:
        """Copy a file to the client's OneDrive synced folder."""
        if not os.path.exists(source_path):
            raise FileNotFoundError(f"Source file not found: {source_path}")

        target_dir = os.path.join(onedrive_origin, client_onedrive_folder)
        os.makedirs(target_dir, exist_ok=True)

        filename = os.path.basename(source_path)
        target_path = os.path.join(target_dir, filename)

        shutil.copy2(source_path, target_path)
        logger.info("Copied to OneDrive", source=source_path, target=target_path)

        return target_path

    def build_sharepoint_url(
        self,
        onedrive_folder: str,
        fy_year: int,
        local_prefix: str,
        sharepoint_base_url: str,
        filename: str | None = None,
    ) -> str | None:
        """Derive a SharePoint web URL from an OneDrive folder path."""
        if not sharepoint_base_url:
            return None

        if onedrive_folder.startswith(local_prefix):
            relative = onedrive_folder[len(local_prefix):].lstrip('\\/')
        else:
            relative = onedrive_folder

        encoded_path = '/'.join(
            urllib.parse.quote(seg, safe='')
            for seg in relative.replace('\\', '/').split('/')
            if seg
        )

        base = sharepoint_base_url.rstrip('/')
        if encoded_path:
            url = f"{base}/{encoded_path}/FY%20{fy_year}"
        else:
            url = f"{base}/FY%20{fy_year}"

        if filename:
            url = f"{url}/{urllib.parse.quote(filename, safe='')}"
        return url

    def cleanup_job_files(self, file_paths: List[str]) -> dict:
        """Delete a list of files."""
        deleted = 0
        cleanup_errors = []

        for path in file_paths:
            try:
                if os.path.exists(path):
                    os.remove(path)
                    deleted += 1
            except Exception as e:
                cleanup_errors.append(f"{os.path.basename(path)}: {str(e)}")

        return {"deleted": deleted, "errors": cleanup_errors}

    def validate_excel_file(self, filepath: str) -> bool:
        """Validate that a file appears to be a valid Excel file."""
        if not os.path.exists(filepath):
            return False

        size = os.stat(filepath).st_size
        if size < 1000:
            return False

        if not filepath.lower().endswith(('.xlsx', '.xls')):
            return False

        try:
            with open(filepath, 'rb') as f:
                header = f.read(4)
                if header[:2] != b'PK':
                    return False
        except Exception:
            return False

        return True

    def consolidate_excel_files(
        self,
        file_paths: List[str],
        output_filename: str,
        sheet_names: Optional[List[str]] = None,
        sheet_name_map: Optional[dict[str, list[str]]] = None,
    ) -> str:
        """
        Consolidate multiple Excel files into a single file.

        Supports two modes:
        1. sheet_names (legacy): One name per file. Multi-sheet files get "{name}_{original}" naming.
        2. sheet_name_map (new): Maps file_path -> list of target sheet names.
           For multi-sheet source files, target names are applied in order to source sheets.
           This is used by the BAS profile where Activity Statement produces 3 sheets
           mapped to ["GST Summary", "GST Detail", "BAS field"].

        Args:
            file_paths: List of paths to Excel files to consolidate
            output_filename: Name for the consolidated output file
            sheet_names: Optional list of custom sheet names (one per file) — legacy mode
            sheet_name_map: Optional dict mapping file_path to list of target sheet names

        Returns:
            Path to the consolidated file
        """
        if not file_paths:
            raise ValueError("No files provided for consolidation")

        for path in file_paths:
            if not os.path.exists(path):
                raise FileNotFoundError(f"File not found: {path}")

        logger.info("Consolidating Excel files",
                   file_count=len(file_paths),
                   output=output_filename)

        output_wb = Workbook()
        default_sheet = output_wb.active

        sheet_index = 0
        for file_idx, file_path in enumerate(file_paths):
            try:
                source_wb = load_workbook(file_path, data_only=False)
                target_names = None

                # Determine target sheet names
                if sheet_name_map and file_path in sheet_name_map:
                    target_names = sheet_name_map[file_path]
                elif sheet_names and file_idx < len(sheet_names):
                    target_names = [sheet_names[file_idx]]

                for ws_idx, source_sheet in enumerate(source_wb.worksheets):
                    # Determine the target sheet name
                    if target_names and ws_idx < len(target_names):
                        new_sheet_name = target_names[ws_idx]
                    elif target_names and len(target_names) == 1 and len(source_wb.worksheets) > 1:
                        # Single target name but multi-sheet source: append original name
                        new_sheet_name = f"{target_names[0]}_{source_sheet.title}"
                    else:
                        # Fallback: use source filename + sheet title
                        base_name = os.path.splitext(os.path.basename(file_path))[0]
                        short_base = base_name[:15] if len(base_name) > 15 else base_name
                        new_sheet_name = f"{short_base}_{source_sheet.title}"

                    new_sheet_name = self._sanitize_sheet_name(new_sheet_name)
                    new_sheet_name = self._make_unique_sheet_name(output_wb, new_sheet_name)

                    new_sheet = output_wb.create_sheet(title=new_sheet_name)
                    self._copy_sheet_data(source_sheet, new_sheet)

                    sheet_index += 1
                    logger.debug("Copied sheet",
                               source=source_sheet.title,
                               target=new_sheet_name)

                source_wb.close()

            except Exception as e:
                logger.error("Error processing file", file=file_path, error=str(e))
                raise

        if sheet_index > 0 and default_sheet.title == "Sheet":
            output_wb.remove(default_sheet)

        output_path = os.path.join(self.download_dir, output_filename)
        output_wb.save(output_path)
        output_wb.close()

        logger.info("Consolidation complete",
                   output_path=output_path,
                   total_sheets=sheet_index)

        return output_path

    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize a string for use as Excel sheet name (max 31 chars)."""
        invalid_chars = '\\/*?:[]'
        result = name
        for char in invalid_chars:
            result = result.replace(char, '_')
        return result[:31]

    def _make_unique_sheet_name(self, workbook: Workbook, name: str) -> str:
        """Ensure sheet name is unique in workbook."""
        existing_names = [sheet.title for sheet in workbook.worksheets]

        if name not in existing_names:
            return name

        counter = 1
        while True:
            suffix = f"_{counter}"
            max_base_len = 31 - len(suffix)
            new_name = f"{name[:max_base_len]}{suffix}"
            if new_name not in existing_names:
                return new_name
            counter += 1

    def _copy_sheet_data(self, source_sheet, target_sheet) -> None:
        """Copy all data and formatting from source sheet to target sheet."""
        from openpyxl.cell.cell import MergedCell

        for row_idx, row in enumerate(source_sheet.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                if isinstance(cell, MergedCell):
                    continue

                target_cell = target_sheet.cell(row=row_idx, column=col_idx)
                target_cell.value = cell.value

                if cell.has_style:
                    target_cell.font = cell.font.copy()
                    target_cell.fill = cell.fill.copy()
                    target_cell.border = cell.border.copy()
                    target_cell.alignment = cell.alignment.copy()
                    target_cell.number_format = cell.number_format

        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))

        for col_letter, col_dim in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[col_letter].width = col_dim.width

        for row_num, row_dim in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[row_num].height = row_dim.height


# Singleton instance
_file_manager: FileManager = None


def get_file_manager() -> FileManager:
    """Get the singleton file manager instance."""
    global _file_manager
    if _file_manager is None:
        _file_manager = FileManager()
    return _file_manager
